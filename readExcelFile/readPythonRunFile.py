import openpyxl
ExcelFileNameList, TestCaseNameList = [], []
def run():
    ExcelFileName, TestCaseName, ExecutionColumnNo = None, None, 1
    wk = openpyxl.load_workbook("A:/PythonTest//PythonRun.xlsx")
    sh = wk['Main']
    rowCount, columnCount = sh.max_row, sh.max_column
    for i in range(1, rowCount + 1):
        for j in range(1, columnCount +1):
            if sh.cell(i, j).value == 'ExcelFileName':
                ExcelFileNameColumnNo = j;
            if sh.cell(i, j).value == 'TestCaseName':
                TestCaseNameColumnNo = j;
            if sh.cell(i, j).value == 'Execution':
                ExecutionColumnNo = j;
            if sh.cell(i, ExecutionColumnNo).value == 'Yes':
                ExcelFileNameList.append(sh.cell(i, ExcelFileNameColumnNo).value)
                TestCaseNameList.append(sh.cell(i, TestCaseNameColumnNo).value)
                break

def scenarioFile():
    BussinessFlowList=[]
    TestCaseNameColumnNo,n=1,0
    for m in ExcelFileNameList:
        wk=openpyxl.load_workbook("A:/PythonTest/"+m+".xlsx")
        sh=wk['BussinessFlow']
        rowCountNum,columnCountNum=sh.max_row,sh.max_column
        for i in range(1,rowCountNum+1):
            if TestCaseNameList[ExcelFileNameList.index(m)]==sh.cell(i,1).value:
                for j in range(2,columnCountNum+1):
                    BussinessFlowList.append(sh.cell(i,j).value);
                break
        print(BussinessFlowList)

run()
if len(ExcelFileNameList)==0:
    print("Test Case execution status for any test case is not Yes")
else :
    scenarioFile()



